import json
import os
import openpyxl
import shutil


def populate_from_json(source_json_file, target_file):
    if os.path.exists(target_file):
        if target_file.endswith(".xlsx"):
            populate_xlsx_file_from_json(source_json_file, target_file)
        else:
            populate_text_file_from_json(source_json_file, target_file)
    else:
        print("File {} not found, cannot process".format(target_file))


def populate_text_file_from_json(source_json_file, target_file):
    with open(source_json_file) as src:
        declaration_dict = json.load(src)
        for key in declaration_dict.keys():  # flatten a dictionary
            if type(declaration_dict[key]) == list:
                declaration_dict[key] = " ".join(declaration_dict[key])

    result_file = target_file.split(".")[0] + "_result." + target_file.split(".")[-1]
    f = open(result_file, "w")
    try:
        with open(target_file) as tf:
            for line in tf.readlines():
                res_line = line
                for key in declaration_dict.keys():
                    res_line = res_line.replace(key, declaration_dict[key])
                f.write(res_line)
    except (OSError, IOError) as e:
        print(e.message)
        print("Cannot process file {}".format(target_file))
    finally:
        f.close()


def populate_xlsx_file_from_json(source_json_file, target_file):
    with open(source_json_file) as src:
        declaration_dict = json.load(src)
    result_file = target_file.split(".")[0] + "_result." + target_file.split(".")[-1]
    shutil.copy(target_file, result_file)
    try:
        xls_file = openpyxl.load_workbook(result_file)
        sheet1 = xls_file.get_sheet_by_name('r1-6')
        MAX_ROWS = 300
        MAX_COLUMNS = 150
        for num_r in range(1, MAX_ROWS):
            for num_c in range(1, MAX_COLUMNS):
                if sheet1.cell(num_r, num_c).value is not None:
                    for key in declaration_dict.keys():
                        if sheet1.cell(num_r, num_c).value == key:
                            if type(declaration_dict[key]) == list:
                                declaration_dict[key] = " ".join(declaration_dict[key])
                            sheet1.cell(num_r, num_c).value = declaration_dict[key]
        xls_file.save(result_file)
    except (OSError, IOError) as e:
        print(e.message)
        print("Cannot process file {}".format(target_file))
    finally:
        xls_file.close()


if __name__ == "__main__":
    source = os.path.join("..", "Documents", "all_fields.json")
    target = os.path.join("..", "Documents", "declarac.xlsx")
    populate_from_json(source, target)
